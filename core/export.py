"""
Универсальный сервис экспорта данных в CSV / XLSX / PDF.

Использование:

    from core.export import ExportColumn, dispatch_export

    columns = [
        ExportColumn("SKU", lambda obj: obj.product.internal_sku),
        ExportColumn("Кол-во", lambda obj: obj.quantity),
    ]

    if request.GET.get("export") in {"csv", "xlsx", "pdf"}:
        return dispatch_export(request, qs, columns,
                               filename="movements", title="Журнал движений")

Сервис не зависит от конкретной модели — работает с любой iterable / queryset.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any
from collections.abc import Callable, Iterable, Sequence

from django.http import HttpResponse
from django.utils import timezone


@dataclass(frozen=True)
class ExportColumn:
    """Описание колонки для экспорта."""
    header: str
    getter: Callable[[Any], Any]


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _row(obj: Any, columns: Sequence[ExportColumn]) -> list[str]:
    return [_safe_str(col.getter(obj)) for col in columns]


def _filename(base: str, ext: str) -> str:
    ts = timezone.now().strftime("%Y%m%d_%H%M%S")
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in base)
    return f"{safe}_{ts}.{ext}"


# ─────────────────────────── CSV ───────────────────────────

def export_csv(
    rows: Iterable[Any],
    columns: Sequence[ExportColumn],
    filename: str = "export",
) -> HttpResponse:
    """CSV с разделителем ';' и BOM (UTF-8) для совместимости с Excel."""
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="{_filename(filename, "csv")}"'
    writer = csv.writer(response, delimiter=";")
    writer.writerow([col.header for col in columns])
    for obj in rows:
        writer.writerow(_row(obj, columns))
    return response


# ─────────────────────────── XLSX ──────────────────────────

def export_xlsx(
    rows: Iterable[Any],
    columns: Sequence[ExportColumn],
    filename: str = "export",
    sheet_name: str = "Export",
) -> HttpResponse:
    """XLSX через openpyxl: шапка жирная, фильтры, авто-ширина."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Export")[:31]  # XLSX limit 31 char

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # шапка
    headers = [col.header for col in columns]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # данные + расчёт ширины
    widths = [len(h) for h in headers]
    for obj in rows:
        values = _row(obj, columns)
        ws.append(values)
        for i, v in enumerate(values):
            widths[i] = min(60, max(widths[i], len(v)))

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename(filename, "xlsx")}"'
    return response


# ─────────────────────────── PDF ───────────────────────────

def export_pdf(
    rows: Iterable[Any],
    columns: Sequence[ExportColumn],
    filename: str = "export",
    title: str = "Отчёт",
) -> HttpResponse:
    """
    PDF через reportlab. Альбомная ориентация, поддержка кириллицы
    через встроенный шрифт DejaVuSans (если есть) или Helvetica.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Шрифт с кириллицей
    font_name = "Helvetica"
    try:
        # стандартный путь в Linux/Win — попытка регистрации
        candidates = [
            "C:/Windows/Fonts/arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/Library/Fonts/Arial.ttf",
        ]
        for path in candidates:
            try:
                pdfmetrics.registerFont(TTFont("WMSFont", path))
                font_name = "WMSFont"
                break
            except Exception:
                continue
    except Exception:
        pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "WMSTitle", parent=styles["Title"], fontName=font_name, fontSize=14
    )
    meta_style = ParagraphStyle(
        "WMSMeta", parent=styles["Normal"], fontName=font_name, fontSize=8,
        textColor=colors.grey,
    )

    elements: list = [
        Paragraph(title, title_style),
        Paragraph(
            timezone.localtime(timezone.now()).strftime("Сформировано: %d.%m.%Y %H:%M"),
            meta_style,
        ),
        Spacer(1, 4 * mm),
    ]

    data: list[list[str]] = [[col.header for col in columns]]
    for obj in rows:
        data.append(_row(obj, columns))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)

    doc.build(elements)

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_filename(filename, "pdf")}"'
    return response


# ───────────────────── единая точка входа ────────────────────

def dispatch_export(
    request,
    rows: Iterable[Any],
    columns: Sequence[ExportColumn],
    filename: str = "export",
    title: str | None = None,
) -> HttpResponse | None:
    """
    Если в request.GET есть export=csv|xlsx|pdf — возвращает соответствующий
    HttpResponse, иначе None (вызвавший код продолжит обычный рендер).
    """
    fmt = (request.GET.get("export") or "").strip().lower()
    if fmt == "csv":
        return export_csv(rows, columns, filename=filename)
    if fmt == "xlsx":
        return export_xlsx(rows, columns, filename=filename, sheet_name=title or filename)
    if fmt == "pdf":
        return export_pdf(rows, columns, filename=filename, title=title or filename)
    return None
