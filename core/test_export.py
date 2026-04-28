"""
Тесты универсального сервиса экспорта (CSV / XLSX / PDF).
"""
import io
from dataclasses import dataclass

from django.test import RequestFactory, TestCase

from core.export import (
    ExportColumn,
    dispatch_export,
    export_csv,
    export_pdf,
    export_xlsx,
)


@dataclass
class _Row:
    sku: str
    name: str
    qty: int


_SAMPLE = [
    _Row("A-1", "Изделие А", 10),
    _Row("B-2", "Изделие Б", 0),
    _Row("C-3", "Изделие В", -5),
]
_COLS = [
    ExportColumn("SKU", lambda r: r.sku),
    ExportColumn("Наименование", lambda r: r.name),
    ExportColumn("Кол-во", lambda r: r.qty),
]


class TestCsv(TestCase):
    def test_csv_has_bom_and_headers(self):
        resp = export_csv(_SAMPLE, _COLS, filename="t")
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8-sig")
        self.assertIn("SKU;Наименование;Кол-во", body)
        self.assertIn("A-1;Изделие А;10", body)
        self.assertEqual(resp["Content-Type"], "text/csv; charset=utf-8-sig")
        self.assertIn(".csv", resp["Content-Disposition"])

    def test_csv_handles_negative_and_zero(self):
        resp = export_csv(_SAMPLE, _COLS, filename="t")
        body = resp.content.decode("utf-8-sig")
        self.assertIn("B-2;Изделие Б;0", body)
        self.assertIn("C-3;Изделие В;-5", body)

    def test_csv_empty(self):
        resp = export_csv([], _COLS, filename="t")
        body = resp.content.decode("utf-8-sig")
        self.assertEqual(body.strip(), "SKU;Наименование;Кол-во")


class TestXlsx(TestCase):
    def test_xlsx_is_valid_openpyxl_file(self):
        from openpyxl import load_workbook

        resp = export_xlsx(_SAMPLE, _COLS, filename="t", sheet_name="Sample")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, "Sample")
        self.assertEqual([c.value for c in ws[1]], ["SKU", "Наименование", "Кол-во"])
        self.assertEqual(ws.cell(row=2, column=1).value, "A-1")
        self.assertEqual(ws.cell(row=2, column=3).value, "10")
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_xlsx_long_sheet_name_truncated(self):
        from openpyxl import load_workbook
        long_name = "A" * 50
        resp = export_xlsx(_SAMPLE, _COLS, sheet_name=long_name)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertLessEqual(len(wb.active.title), 31)


class TestPdf(TestCase):
    def test_pdf_starts_with_pdf_magic(self):
        resp = export_pdf(_SAMPLE, _COLS, filename="t", title="Тестовый отчёт")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content.startswith(b"%PDF-"))
        self.assertGreater(len(resp.content), 1000)

    def test_pdf_empty_dataset(self):
        resp = export_pdf([], _COLS, filename="t", title="Empty")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.content.startswith(b"%PDF-"))


class TestDispatch(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_dispatch_returns_none_without_export(self):
        req = self.factory.get("/x/")
        self.assertIsNone(dispatch_export(req, _SAMPLE, _COLS))

    def test_dispatch_csv(self):
        req = self.factory.get("/x/?export=csv")
        resp = dispatch_export(req, _SAMPLE, _COLS, filename="x")
        self.assertEqual(resp["Content-Type"], "text/csv; charset=utf-8-sig")

    def test_dispatch_xlsx(self):
        req = self.factory.get("/x/?export=xlsx")
        resp = dispatch_export(req, _SAMPLE, _COLS, filename="x")
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_dispatch_pdf(self):
        req = self.factory.get("/x/?export=pdf")
        resp = dispatch_export(req, _SAMPLE, _COLS, filename="x", title="t")
        self.assertEqual(resp["Content-Type"], "application/pdf")

    def test_dispatch_unknown_format(self):
        req = self.factory.get("/x/?export=ZZZ")
        self.assertIsNone(dispatch_export(req, _SAMPLE, _COLS))
